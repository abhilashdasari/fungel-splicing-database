#!/usr/bin/env python3
"""Export workbook gene-list sheets into GitHub Pages-friendly CSV files."""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
APP_ROOT = Path(__file__).resolve().parents[1]
ANALYSIS_DIR = ROOT / "analysis_pipeline" / "19_splicing_event_analysis"
WORKBOOK = ANALYSIS_DIR / "splicing_comparisons_68_sheets.xlsx"
EVENT_DISTRIBUTION = ANALYSIS_DIR / "01_splicing_event_distribution.csv"
NAME_MAPPING = ANALYSIS_DIR / "comparison_name_mapping.csv"
MASTER_GENE_PRESENCE = ROOT / "analysis_pipeline" / "02_csv_outputs" / "master_gene_level_presence.csv"
OUT_DIR = APP_ROOT / "data" / "gene_lists"
MANIFEST = APP_ROOT / "data" / "gene_list_manifest.json"
UNIQUE_GENE_SETS = APP_ROOT / "data" / "unique_gene_sets.json"

CATEGORY_LABELS = {
    "upregulated_genes": "Upregulated genes",
    "downregulated_genes": "Downregulated genes",
    "Differential.exon.usage": "Differential exon usage",
    "DIfferential.transcript.exxpression": "Differential transcript expression",
    "Differential.intron.excision": "Differential intron excision",
    "Differential.alternative.splicing": "Differential alternative splicing",
}


def slugify(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return value.strip("_") or "comparison"


def read_csv_dict(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    distribution_rows = read_csv_dict(EVENT_DISTRIBUTION)
    comparison_ids = [row["comparison_id"] for row in distribution_rows[:68]]
    distribution_by_id = {row["comparison_id"]: row for row in distribution_rows}
    short_names = {
        row["old_name"]: row["new_name"]
        for row in read_csv_dict(NAME_MAPPING)
        if row.get("old_name") and row.get("new_name")
    }

    workbook = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    if len(workbook.sheetnames) != len(comparison_ids):
        raise RuntimeError(
            f"Workbook has {len(workbook.sheetnames)} sheets, but expected {len(comparison_ids)} comparisons"
        )

    manifest = []
    unique_gene_sets = {comparison_id: [] for comparison_id in comparison_ids}
    master_unique_sets = {comparison_id: set() for comparison_id in comparison_ids}
    comparison_id_set = set(comparison_ids)

    with MASTER_GENE_PRESENCE.open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            comparison_id = row.get("comparison_id", "")
            gene_symbol = row.get("gene_symbol", "").strip()
            if comparison_id in comparison_id_set and gene_symbol and row.get("in_dsg") == "1":
                master_unique_sets[comparison_id].add(gene_symbol)

    for index, (sheet_name, comparison_id) in enumerate(zip(workbook.sheetnames, comparison_ids), start=1):
        worksheet = workbook[sheet_name]
        worksheet.reset_dimensions()
        rows_iter = worksheet.iter_rows(values_only=True)
        headers = [str(value).strip() for value in next(rows_iter)]

        gene_rows = []
        category_counts = {CATEGORY_LABELS.get(header, header): 0 for header in headers}
        seen_genes = set()

        for source_row in rows_iter:
            for header, value in zip(headers, source_row):
                if value is None:
                    continue
                gene_name = str(value).strip()
                if not gene_name:
                    continue

                category = CATEGORY_LABELS.get(header, header)
                category_counts[category] = category_counts.get(category, 0) + 1
                seen_genes.add(gene_name)
                gene_rows.append(
                    {
                        "comparison_id": comparison_id,
                        "comparison_name": short_names.get(comparison_id, comparison_id),
                        "gene_set": category,
                        "gene_name": gene_name,
                    }
                )

        filename = f"{index:02d}_{slugify(comparison_id)}.csv"
        write_csv(
            OUT_DIR / filename,
            gene_rows,
            ["comparison_id", "comparison_name", "gene_set", "gene_name"],
        )

        meta = distribution_by_id.get(comparison_id, {})
        manifest.append(
            {
                "comparison_id": comparison_id,
                "display_name": short_names.get(comparison_id, comparison_id),
                "condition": meta.get("condition", ""),
                "disease_group": meta.get("disease_group_combined") or meta.get("disease_group", ""),
                "filename": f"gene_lists/{filename}",
                "row_count": len(gene_rows),
                "unique_gene_count": len(seen_genes),
                "category_counts": category_counts,
            }
        )
        unique_gene_sets[comparison_id] = sorted(master_unique_sets.get(comparison_id, set()))

    with MANIFEST.open("w", encoding="utf-8") as handle:
        json.dump(manifest, handle, indent=2)
        handle.write("\n")

    with UNIQUE_GENE_SETS.open("w", encoding="utf-8") as handle:
        json.dump(unique_gene_sets, handle, separators=(",", ":"))
        handle.write("\n")

    print(f"Exported {len(manifest)} comparison gene-list files to {OUT_DIR}")
    print(f"Wrote {MANIFEST}")
    print(f"Wrote {UNIQUE_GENE_SETS}")


if __name__ == "__main__":
    main()
